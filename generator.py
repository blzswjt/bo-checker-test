"""
数据建模生成智能体 - 核心生成逻辑
从需求文档(docx)中自动提取业务对象、逻辑实体、业务属性，生成标准化Excel。

模块分区：
  1. 文档解析: parse_docx, extract_docx_images
  2. 图片分析: analyze_images
  3. LLM编排: extract_business_objects, extract_logical_entities, extract_business_attributes
  4. Excel生成: generate_excel
  5. 主管线: run_generation_pipeline (SSE生成器)
"""
import json
import re
import base64
import zipfile
import uuid
from pathlib import Path
from typing import Generator

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from docx.oxml.ns import qn
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from llm import chat_stream, analyze_image, get_model_display_name
from rules import ELEMENT_RULES

# ============================================================
# 1. 文档解析
# ============================================================

def parse_docx(file_path: str) -> dict:
    """
    解析docx文件，提取段落文本（保留标题层级）、表格数据、图片信息。
    返回: {sections, tables, image_count, full_text, structure}
    """
    path = Path(file_path)
    if not path.exists():
        return {"error": "文件不存在"}
    if not path.suffix.lower() == ".docx":
        return {"error": "请上传 .docx 格式文件"}

    try:
        doc = Document(file_path)
    except Exception as e:
        return {"error": f"解析文档失败: {str(e)}"}

    sections = []  # [{level, title, content_preview}]
    full_text_parts = []
    tables_data = []
    image_count = 0

    # 遍历文档body中的所有元素（段落和表格按顺序）
    for element in doc.element.body:
        if element.tag == qn('w:p'):
            para = Paragraph(element, doc)
            text = para.text.strip()
            if not text:
                continue

            # 检测标题层级
            style_name = para.style.name if para.style else ""
            level = 0
            if "Heading" in style_name or "标题" in style_name:
                try:
                    level = int(re.search(r'\d', style_name).group()) if re.search(r'\d', style_name) else 1
                except (AttributeError, ValueError):
                    level = 1

            # 检测图片
            drawings = para._element.findall('.//' + qn('w:drawing'))
            pics = para._element.findall('.//' + qn('w:pict'))
            if drawings or pics:
                image_count += len(drawings) + len(pics)

            if level > 0:
                sections.append({
                    "level": level,
                    "title": text,
                    "content_preview": ""
                })
                full_text_parts.append(f"\n{'#' * level} {text}\n")
            else:
                full_text_parts.append(text)
                # 更新最近section的内容预览
                if sections and len(sections[-1]["content_preview"]) < 200:
                    sections[-1]["content_preview"] += text[:100] + " "

        elif element.tag == qn('w:tbl'):
            table = Table(element, doc)
            table_data = _parse_table(table)
            if table_data:
                tables_data.append(table_data)
                # 将表格内容也加入全文
                full_text_parts.append(_table_to_text(table_data))

    full_text = "\n".join(full_text_parts)

    # 用zip方式准确统计图片数量（段落XML搜索可能遗漏）
    actual_image_count = image_count
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            media_files = [f for f in z.namelist() if f.startswith('word/media/')]
            actual_image_count = max(image_count, len(media_files))
    except Exception:
        pass

    return {
        "sections": sections,
        "tables": tables_data[:30],  # 限制返回表格数量
        "table_count": len(tables_data),
        "image_count": actual_image_count,
        "full_text": full_text,
        "text_length": len(full_text),
        "structure": _build_structure_tree(sections),
    }


def _parse_table(table: Table) -> dict:
    """解析单个表格为结构化数据"""
    rows = []
    for row in table.rows:
        cells = [cell.text.strip() for cell in row.cells]
        rows.append(cells)

    if not rows or len(rows) < 2:
        return None

    # 第一行作为表头
    headers = rows[0]
    data_rows = rows[1:]

    return {
        "headers": headers,
        "rows": data_rows[:100],  # 限制行数
        "row_count": len(data_rows),
        "col_count": len(headers),
    }


def _table_to_text(table_data: dict) -> str:
    """将表格数据转为文本格式（供LLM阅读）"""
    headers = table_data["headers"]
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in table_data["rows"][:50]:  # 限制50行
        # 补齐列数
        padded = row + [""] * (len(headers) - len(row))
        lines.append("| " + " | ".join(padded[:len(headers)]) + " |")
    if table_data["row_count"] > 50:
        lines.append(f"... (共{table_data['row_count']}行)")
    return "\n".join(lines)


def _build_structure_tree(sections: list) -> list:
    """构建文档结构树（简化版）"""
    tree = []
    for s in sections:
        tree.append({
            "level": s["level"],
            "title": s["title"],
        })
    return tree


# ============================================================
# 2. 图片提取与分析
# ============================================================

def extract_docx_images(file_path: str) -> list[dict]:
    """
    从docx中提取所有图片。
    返回: [{name, base64, size, format}]
    """
    images = []
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            media_files = [f for f in z.namelist() if f.startswith('word/media/')]
            for mf in media_files:
                data = z.read(mf)
                name = Path(mf).name
                # 判断格式
                fmt = "png"
                if name.lower().endswith((".jpg", ".jpeg")):
                    fmt = "jpeg"
                elif name.lower().endswith(".gif"):
                    fmt = "gif"
                elif name.lower().endswith(".bmp"):
                    fmt = "bmp"

                images.append({
                    "name": name,
                    "base64": base64.b64encode(data).decode("utf-8"),
                    "size": len(data),
                    "format": fmt,
                })
    except Exception as e:
        return []

    return images


def analyze_images(images: list[dict], model_id: str = None) -> str:
    """
    调用视觉模型分析文档中的图片（架构图、流程图、E-R图等）。
    返回合并的分析文本。
    """
    if not images:
        return ""

    prompt = """请仔细分析这张图片，它来自一份数据建模/产品需求文档。请提取以下信息：

1. 如果是业务架构图：提取L1-L4各层级的名称和层级关系
2. 如果是E-R图/数据模型图：提取所有实体名称、实体间关系（1:1, 1:N, M:N）、主要属性
3. 如果是流程图：提取关键业务节点、流转关系、涉及的业务对象
4. 如果是状态图：提取状态名称和流转条件
5. 如果是原型图/界面截图：提取页面中涉及的数据字段和业务对象
6. 如果是表格截图：提取表格中的所有文字内容

请以结构化文本形式输出提取到的信息，重点标注：
- 业务对象名称
- 逻辑实体名称
- 实体间关系
- 层级归属关系
- 图片中的文字内容

只输出提取到的结构化信息，不要输出无关内容。"""

    results = []
    # 过滤掉太小的图片（小于5KB通常是图标/装饰）
    meaningful_images = [img for img in images if img["size"] > 5 * 1024]
    # 分析所有有意义的图片（最多20张）
    for img in meaningful_images[:20]:
        if img["size"] > 10 * 1024 * 1024:  # 跳过超过10MB的
            continue
        try:
            result = analyze_image(img["base64"], prompt, model_id=model_id)
            if result and len(result.strip()) > 20:
                results.append(f"### 图片: {img['name']}\n{result}")
        except Exception as e:
            results.append(f"### 图片: {img['name']}\n[分析失败: {str(e)}]")

    return "\n\n".join(results)


# ============================================================
# 3. LLM 多步编排
# ============================================================

def _build_rules_text(element_type: str) -> str:
    """构建指定元素类型的规则文本"""
    rules = ELEMENT_RULES.get(element_type)
    if not rules:
        return ""

    parts = [f"## {element_type}定义\n{rules['description']}"]

    parts.append("\n## 识别规则")
    for i, r in enumerate(rules.get("identification", []), 1):
        if r.get("enabled", True) is False:
            continue
        parts.append(f"{i}. 【{r['rule']}】{r['desc']}")
        if r.get("positive"):
            parts.append(f"   正例：{r['positive']}")
        if r.get("negative"):
            parts.append(f"   反例：{r['negative']}")

    parts.append("\n## 命名规则")
    for i, r in enumerate(rules.get("naming", []), 1):
        if r.get("enabled", True) is False:
            continue
        parts.append(f"{i}. 【{r['rule']}】{r['desc']}")
        if r.get("positive"):
            parts.append(f"   正例：{r['positive']}")
        if r.get("negative"):
            parts.append(f"   反例：{r['negative']}")

    if rules.get("definition"):
        parts.append("\n## 定义规则")
        for i, r in enumerate(rules["definition"], 1):
            if r.get("enabled", True) is False:
                continue
            parts.append(f"{i}. 【{r['rule']}】{r['desc']}")
            if r.get("positive"):
                parts.append(f"   正例：{r['positive']}")

    if rules.get("not_examples"):
        parts.append("\n## 常见不是的情况")
        for ex in rules["not_examples"]:
            parts.append(f"- {ex}")

    return "\n".join(parts)


def _extract_json_from_response(text: str) -> list | dict | None:
    """从LLM响应中提取JSON数据"""
    # 方案1: ```json 代码块
    m = re.search(r'```json\s*([\s\S]*?)\s*```', text)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # 方案2: 直接找 [ 或 { 开头
    for start_char, end_char in [('[', ']'), ('{', '}')]:
        start = text.find(start_char)
        if start >= 0:
            # 从后往前找配对的结束符
            end = text.rfind(end_char)
            if end > start:
                try:
                    return json.loads(text[start:end + 1])
                except json.JSONDecodeError:
                    pass

    return None


def extract_business_objects(doc_context: str, model_id: str = None) -> Generator:
    """
    SSE生成器：从文档上下文中提取业务对象清单。
    yield SSE事件dict。
    """
    rules_text = _build_rules_text("业务对象")

    system_prompt = """你是一位资深数据架构师，擅长企业数据建模。请根据提供的需求文档内容，识别并提取所有业务对象。

严格按照以下规则进行识别和命名：
""" + rules_text + """

## 输出要求

请分析文档内容，识别所有符合业务对象定义的事物，然后输出JSON数组。

每个业务对象包含：
- l1: 主题域分类（对应业务架构L1，如"线索到回款"）
- l2: 主题域分组（对应业务架构L2）
- l3: 主题域（对应业务架构L3）
- code: 业务对象编码（格式：YWDX-NNNNNN，从000001开始递增）
- da_code: 数据架构编码（格式：DA_NNNNNNNN）
- name_cn: 中文名称（必须是名词，符合命名规则）
- name_en: 英文名称（PascalCase）
- definition: 定义说明（包含目的、定义、范围，200字以上）
- data_class: 数据分类（"主数据"或"事务数据"）
- entities: 该业务对象下的逻辑实体名称列表（初步识别）
- source: 来源章节（从文档哪个章节/段落提取的，如"7.1 整体业务框架"、8.1 L4业务对象设计"）

先输出你的分析思考过程，最后输出JSON：
```json
[{"l1":"...", "l2":"...", "l3":"...", "code":"YWDX-000001", "da_code":"DA_03010101", "name_cn":"...", "name_en":"...", "definition":"...", "data_class":"事务数据", "entities":["实体1","实体2"], "source":"章节名"}]
```"""

    user_prompt = f"""## 需求文档内容

{doc_context[:15000]}

---

请从以上文档中识别所有业务对象，严格按照规则进行判断和命名。"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ]

    full_response = ""
    for token in chat_stream(messages, temperature=0.2, model_id=model_id):
        full_response += token
        yield {"type": "thinking", "phase": "business_objects", "token": token}

    # 解析结果
    parsed = _extract_json_from_response(full_response)
    if parsed and isinstance(parsed, list):
        yield {"type": "bo_result", "data": parsed}
    else:
        yield {"type": "error", "phase": "business_objects", "message": "无法解析业务对象结果，请重试"}
        yield {"type": "bo_result", "data": []}


def extract_logical_entities(doc_context: str, business_objects: list, model_id: str = None) -> Generator:
    """
    SSE生成器：为每个业务对象提取逻辑实体。
    """
    rules_text = _build_rules_text("逻辑实体")

    bo_summary = ""
    for bo in business_objects:
        entities_hint = "、".join(bo.get("entities", [])) if bo.get("entities") else "待识别"
        bo_summary += f"- {bo['name_cn']}（{bo.get('code', '')}）：初步实体=[{entities_hint}]\n"

    system_prompt = """你是一位资深数据架构师，擅长企业数据建模。请根据需求文档和已识别的业务对象，为每个业务对象设计逻辑实体。

严格按照以下规则：
""" + rules_text + """

## 输出要求

为每个业务对象设计其下属的逻辑实体，输出JSON数组。

每个逻辑实体包含：
- bo_name: 所属业务对象中文名称
- bo_code: 所属业务对象编码
- entity_code: 逻辑实体编码（格式：DA_业务对象编码后4位+NN，如DA_0301010101）
- name_cn: 中文名称（名词，以下层加前缀规则命名）
- name_en: 英文名称
- definition: 定义（包含目的、定义、范围）
- source: 来源章节（从文档哪个章节提取的，如"8.2 L5-L6逻辑实体设计"、10.1.2 注册合同"）

注意：
1. 每个业务对象有且只有一个主逻辑实体（通常以"XX基本信息"或"XX头"命名）
2. 剔除技术数据（接口表、任务表等）和衍生数据（宽表、汇总表）
3. 属性较多时按高内聚原则拆分为多个子实体
4. 关系实体命名格式为"实体1和实体2关系"

先输出分析思考，最后输出JSON：
```json
[{"bo_name":"...", "bo_code":"...", "entity_code":"...", "name_cn":"...", "name_en":"...", "definition":"...", "source":"章节名"}]
```"""

    user_prompt = f"""## 已识别的业务对象
{bo_summary}

## 需求文档内容（节选）
{doc_context[:12000]}

---

请为每个业务对象设计逻辑实体。"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ]

    full_response = ""
    for token in chat_stream(messages, temperature=0.2, model_id=model_id):
        full_response += token
        yield {"type": "thinking", "phase": "entities", "token": token}

    parsed = _extract_json_from_response(full_response)
    if parsed and isinstance(parsed, list):
        yield {"type": "entity_result", "data": parsed}
    else:
        yield {"type": "error", "phase": "entities", "message": "无法解析逻辑实体结果"}
        yield {"type": "entity_result", "data": []}


def extract_business_attributes(doc_context: str, entities: list, model_id: str = None) -> Generator:
    """
    SSE生成器：为每个逻辑实体提取业务属性。
    按业务对象分批处理。
    """
    rules_text = _build_rules_text("业务属性")

    # 按业务对象分组实体
    bo_groups = {}
    for ent in entities:
        bo_name = ent.get("bo_name", "未知")
        if bo_name not in bo_groups:
            bo_groups[bo_name] = []
        bo_groups[bo_name].append(ent)

    all_attributes = []
    total_bos = len(bo_groups)

    for idx, (bo_name, bo_entities) in enumerate(bo_groups.items()):
        yield {"type": "progress", "phase": "attributes",
               "message": f"正在提取 [{bo_name}] 的业务属性 ({idx+1}/{total_bos})",
               "current": idx, "total": total_bos}

        entity_names = "\n".join(f"- {e['name_cn']}（{e.get('definition', '')[:80]}）" for e in bo_entities)

        system_prompt = f"""你是一位资深数据架构师。请为业务对象「{bo_name}」下的逻辑实体设计业务属性。

严格按照以下规则：
{rules_text}

## 输出要求

为每个逻辑实体设计其包含的业务属性，输出JSON数组。

每个属性包含：
- entity: 所属逻辑实体中文名称
- attr_cn: 属性中文名称（业务词汇、见名知义、词汇简练）
- attr_en: 属性英文名称（snake_case）
- type: 数据类型（VARCHAR/BIGINT/DECIMAL/DATETIME/TEXT/TINYINT等）
- length: 长度（如256、18,2等，数值型可留空）
- required: 是否必填（Y/N）
- logic: 字段逻辑说明（数据来源、计算规则、取值范围等）
- source: 来源章节（从文档哪个章节/表格提取的，如"10.1.2 注册合同-交易要素"）

注意：
1. 剔除技术字段（ID主键、创建人、修改人、删除标记、租户ID等）
2. 每个属性必须是最小业务语义单元（原子性）
3. 属性名称用正式业务词汇，不用口语化表达
4. 外键引用字段保留业务含义名称（如"客户编码"而非"customer_id"）

先简要分析，最后输出JSON：
```json
[{{"entity":"...", "attr_cn":"...", "attr_en":"...", "type":"VARCHAR", "length":"256", "required":"Y", "logic":"...", "source":"章节名"}}]
```"""

        user_prompt = f"""## 业务对象：{bo_name}

## 该对象下的逻辑实体：
{entity_names}

## 需求文档相关内容（节选）
{doc_context[:10000]}

---

请为以上每个逻辑实体设计完整的业务属性列表。"""

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]

        full_response = ""
        for token in chat_stream(messages, temperature=0.2, model_id=model_id):
            full_response += token
            yield {"type": "thinking", "phase": "attributes", "token": token}

        parsed = _extract_json_from_response(full_response)
        if parsed and isinstance(parsed, list):
            all_attributes.extend(parsed)
            yield {"type": "attr_batch_result", "bo_name": bo_name, "data": parsed}

    yield {"type": "attr_result", "data": all_attributes}


# ============================================================
# 4. Excel 生成
# ============================================================

def generate_excel(business_objects: list, entities: list, attributes: list, output_path: str) -> str:
    """
    生成标准格式的Excel文件（三张表）。
    返回输出文件路径。
    """
    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ---- Sheet1: 业务对象清单 ----
    ws1 = wb.active
    ws1.title = "业务对象清单"
    headers1 = ["L1-主题域分类", "L2-主题域分组", "L3-主题域",
                "*业务对象编码", "*业务对象编码", "*业务对象中文名称",
                "*业务对象英文名称", "定义说明", "数据分类识别（主数据、事务数据）",
                "*逻辑实体中文名称", "来源章节"]

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    row_idx = 2
    for bo in business_objects:
        # 找到该业务对象关联的逻辑实体
        bo_entities = [e for e in entities if e.get("bo_name") == bo.get("name_cn")]
        entity_names = [e.get("name_cn", "") for e in bo_entities]

        if not entity_names:
            entity_names = bo.get("entities", [""])
        if not entity_names:
            entity_names = [""]

        start_row = row_idx
        num_rows = len(entity_names)

        # 写入每行的逻辑实体名称（第10列）
        for i, ent_name in enumerate(entity_names):
            ws1.cell(row=row_idx, column=10, value=ent_name)
            for col in range(1, 12):
                ws1.cell(row=row_idx, column=col).border = thin_border
                ws1.cell(row=row_idx, column=col).alignment = Alignment(vertical='center', wrap_text=True)
            row_idx += 1

        # 写入业务对象信息并合并单元格（前9列 + 第11列来源章节）
        bo_values = [
            bo.get("l1", ""), bo.get("l2", ""), bo.get("l3", ""),
            bo.get("code", ""), bo.get("da_code", ""), bo.get("name_cn", ""),
            bo.get("name_en", ""), bo.get("definition", ""), bo.get("data_class", "")
        ]
        for col, val in enumerate(bo_values, 1):
            ws1.cell(row=start_row, column=col, value=val)
            # 多行时合并单元格
            if num_rows > 1:
                ws1.merge_cells(
                    start_row=start_row, start_column=col,
                    end_row=start_row + num_rows - 1, end_column=col
                )
            ws1.cell(row=start_row, column=col).alignment = Alignment(
                vertical='center', horizontal='center' if col <= 5 else 'left', wrap_text=True
            )

        # 来源章节（第11列），同样合并
        source_val = bo.get("source", "")
        ws1.cell(row=start_row, column=11, value=source_val)
        if num_rows > 1:
            ws1.merge_cells(
                start_row=start_row, start_column=11,
                end_row=start_row + num_rows - 1, end_column=11
            )
        ws1.cell(row=start_row, column=11).alignment = Alignment(vertical='center', wrap_text=True)

    # 设置列宽
    col_widths1 = [14, 16, 16, 14, 14, 16, 16, 50, 20, 20, 24]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # ---- Sheet2: 逻辑数据实体清单 ----
    ws2 = wb.create_sheet("逻辑数据实体清单")
    headers2 = ["*业务对象名称", "*业务对象编码", "*逻辑实体编码",
                "*逻辑实体中文名称", "*逻辑实体英文名称", "*逻辑实体定义", "来源章节"]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_i, ent in enumerate(entities, 2):
        ws2.cell(row=row_i, column=1, value=ent.get("bo_name", ""))
        ws2.cell(row=row_i, column=2, value=ent.get("bo_code", ""))
        ws2.cell(row=row_i, column=3, value=ent.get("entity_code", ""))
        ws2.cell(row=row_i, column=4, value=ent.get("name_cn", ""))
        ws2.cell(row=row_i, column=5, value=ent.get("name_en", ""))
        ws2.cell(row=row_i, column=6, value=ent.get("definition", ""))
        ws2.cell(row=row_i, column=7, value=ent.get("source", ""))
        for col in range(1, 8):
            ws2.cell(row=row_i, column=col).border = thin_border
            ws2.cell(row=row_i, column=col).alignment = Alignment(vertical='top', wrap_text=True)

    col_widths2 = [16, 14, 16, 20, 24, 60, 24]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ---- Sheet3: 逻辑数据实体属性清单 ----
    ws3 = wb.create_sheet("逻辑数据实体属性清单")
    headers3 = ["逻辑实体", "属性中文名称", "属性英文名称", "类型", "长度", "是否必填", "逻辑", "来源章节"]

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 按逻辑实体分组，添加分组标题行
    current_entity = None
    row_i = 2
    for attr in attributes:
        ent_name = attr.get("entity", "")
        # 如果切换了逻辑实体，添加分组标题
        if ent_name != current_entity:
            current_entity = ent_name
            # 添加分组标题行（合并单元格效果）
            for col in range(1, 9):
                cell = ws3.cell(row=row_i, column=col, value=ent_name if col == 1 else "")
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                cell.border = thin_border
            row_i += 1

        ws3.cell(row=row_i, column=1, value=ent_name)
        ws3.cell(row=row_i, column=2, value=attr.get("attr_cn", ""))
        ws3.cell(row=row_i, column=3, value=attr.get("attr_en", ""))
        ws3.cell(row=row_i, column=4, value=attr.get("type", ""))
        ws3.cell(row=row_i, column=5, value=str(attr.get("length", "")))
        ws3.cell(row=row_i, column=6, value=attr.get("required", ""))
        ws3.cell(row=row_i, column=7, value=attr.get("logic", ""))
        ws3.cell(row=row_i, column=8, value=attr.get("source", ""))
        for col in range(1, 9):
            ws3.cell(row=row_i, column=col).border = thin_border
            ws3.cell(row=row_i, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        row_i += 1

    col_widths3 = [18, 18, 22, 14, 10, 10, 50, 24]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # 保存
    wb.save(output_path)
    return output_path


# ============================================================
# 5. 主管线（SSE生成器）
# ============================================================

def run_generation_pipeline(file_path: str, model_id: str = None,
                            vision_model_id: str = None,
                            analyze_img: bool = True) -> Generator:
    """
    完整生成管线（SSE生成器）。
    步骤：解析文档 → 分析图片 → 提取业务对象 → 提取逻辑实体 → 提取属性 → 生成Excel
    """
    file_id = str(uuid.uuid4())[:8]

    # Phase 1: 解析文档
    yield {"type": "phase", "phase": "parse", "message": "正在解析文档..."}
    parsed = parse_docx(file_path)
    if "error" in parsed:
        yield {"type": "error", "message": parsed["error"]}
        return

    yield {"type": "parse_done", "sections": len(parsed.get("sections", [])),
           "tables": parsed.get("table_count", 0),
           "images": parsed.get("image_count", 0),
           "text_length": parsed.get("text_length", 0)}

    doc_context = parsed.get("full_text", "")

    # Phase 2: 分析图片（可选）
    image_analysis = ""
    if analyze_img and parsed.get("image_count", 0) > 0:
        yield {"type": "phase", "phase": "images", "message": f"正在分析 {parsed['image_count']} 张图片..."}
        images = extract_docx_images(file_path)
        if images:
            image_analysis = analyze_images(images, model_id=vision_model_id)
            if image_analysis:
                doc_context += "\n\n## 图片分析结果（来自视觉模型）\n" + image_analysis
                yield {"type": "images_done", "count": len(images), "analysis_length": len(image_analysis)}
            else:
                yield {"type": "images_done", "count": len(images), "analysis_length": 0}
    else:
        yield {"type": "phase", "phase": "images", "message": "跳过图片分析"}

    # Phase 3: 提取业务对象
    yield {"type": "phase", "phase": "business_objects", "message": "正在提取业务对象..."}
    business_objects = []
    for event in extract_business_objects(doc_context, model_id=model_id):
        yield event
        if event.get("type") == "bo_result":
            business_objects = event.get("data", [])

    if not business_objects:
        yield {"type": "error", "message": "未能提取到业务对象，流程终止"}
        return

    yield {"type": "phase_done", "phase": "business_objects",
           "count": len(business_objects),
           "message": f"已识别 {len(business_objects)} 个业务对象"}

    # Phase 4: 提取逻辑实体
    yield {"type": "phase", "phase": "entities", "message": "正在提取逻辑实体..."}
    entities = []
    for event in extract_logical_entities(doc_context, business_objects, model_id=model_id):
        yield event
        if event.get("type") == "entity_result":
            entities = event.get("data", [])

    yield {"type": "phase_done", "phase": "entities",
           "count": len(entities),
           "message": f"已识别 {len(entities)} 个逻辑实体"}

    # Phase 5: 提取业务属性
    yield {"type": "phase", "phase": "attributes", "message": "正在提取业务属性..."}
    attributes = []
    for event in extract_business_attributes(doc_context, entities, model_id=model_id):
        yield event
        if event.get("type") == "attr_result":
            attributes = event.get("data", [])

    yield {"type": "phase_done", "phase": "attributes",
           "count": len(attributes),
           "message": f"已识别 {len(attributes)} 个业务属性"}

    # Phase 6: 生成Excel
    yield {"type": "phase", "phase": "excel", "message": "正在生成Excel文件..."}
    output_dir = Path(__file__).parent / "uploads"
    output_dir.mkdir(exist_ok=True)
    output_path = str(output_dir / f"gen_{file_id}_数据建模清单.xlsx")

    try:
        generate_excel(business_objects, entities, attributes, output_path)
    except Exception as e:
        yield {"type": "error", "message": f"Excel生成失败: {str(e)}"}
        return

    # 完成
    yield {
        "type": "done",
        "file_id": file_id,
        "file_name": f"数据建模清单_{file_id}.xlsx",
        "download_path": f"gen_{file_id}_数据建模清单.xlsx",
        "summary": {
            "business_objects": len(business_objects),
            "entities": len(entities),
            "attributes": len(attributes),
        }
    }
